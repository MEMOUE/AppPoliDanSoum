from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Sum, Count, Q, F, Avg
from django.forms import formset_factory
from django.db import transaction
from django.http import HttpResponse, JsonResponse
from .models import (
    ProcesVerbal, ResultatCandidat, BureauVote,
    CentreVote, SousPrefecture, User, Departement
)
from .forms import LoginForm, ProcesVerbalForm, ResultatCandidatForm, ResultatCandidatFormSet


# ========================================
# VUES EXISTANTES (inchangées)
# ========================================

def home(request):
    """Page d'accueil"""
    context = {
        'total_bureaux': BureauVote.objects.count(),
        'total_centres': CentreVote.objects.count(),
        'total_sous_prefectures': SousPrefecture.objects.count(),
        'resultats_saisis': ProcesVerbal.objects.count(),
        'total_candidats': User.objects.filter(role='candidat').count(),
    }
    return render(request, "home.html", context)


def login_view(request):
    """Vue de connexion"""
    if request.user.is_authenticated:
        if request.user.role == 'candidat':
            return redirect('dashboard_general')
        elif request.user.role == 'representant':
            return redirect('saisie_resultat')
        return redirect('home')

    if request.method == 'POST':
        form = LoginForm(request, data=request.POST)
        if form.is_valid():
            user = form.get_user()
            login(request, user)
            messages.success(request, f'Bienvenue {user.get_full_name()}!')

            if user.role == 'candidat':
                return redirect('dashboard_general')
            elif user.role == 'representant':
                return redirect('saisie_resultat')

            return redirect('home')
        else:
            messages.error(request, 'Nom d\'utilisateur ou mot de passe incorrect.')
    else:
        form = LoginForm()

    return render(request, 'login.html', {'form': form})


@login_required
def logout_view(request):
    """Vue de déconnexion"""
    logout(request)
    messages.info(request, 'Vous avez été déconnecté avec succès.')
    return redirect('home')


@login_required
def saisie_resultat(request):
    """Vue pour la saisie complète du procès-verbal par les représentants - VERSION CORRIGÉE"""
    # Vérifier que l'utilisateur est un représentant
    if request.user.role != 'representant':
        messages.error(request, 'Accès non autorisé. Cette page est réservée aux représentants.')
        return redirect('home')

    # Vérifier que le représentant a un bureau affecté
    if not request.user.bureau_vote:
        return render(request, 'saisie_resultat.html', {'no_bureau': True})

    bureau = request.user.bureau_vote

    # Récupérer tous les candidats
    candidats = User.objects.filter(role='candidat').order_by('numero_candidat', 'first_name')

    if not candidats.exists():
        messages.error(request, 'Aucun candidat n\'est enregistré dans le système.')
        return redirect('home')

    # Récupérer le PV existant s'il existe
    try:
        pv_existant = ProcesVerbal.objects.get(bureau_vote=bureau)
    except ProcesVerbal.DoesNotExist:
        pv_existant = None

    # IMPORTANT : Définir le formset UNE SEULE FOIS avec le bon nombre
    # On utilise extra=len(candidats) pour avoir exactement le bon nombre de formulaires
    ResultatFormSet = formset_factory(
        ResultatCandidatForm,
        formset=ResultatCandidatFormSet,
        extra=len(candidats),
        max_num=len(candidats),
        min_num=len(candidats),
        validate_max=True,
        validate_min=True
    )

    if request.method == 'POST':
        pv_form = ProcesVerbalForm(
            request.POST,
            request.FILES,
            instance=pv_existant,
            bureau_vote=bureau
        )

        # Récupérer le nombre d'inscrits du formulaire
        nombre_inscrits_saisi = request.POST.get('nombre_inscrits', 0)
        try:
            nombre_inscrits_saisi = int(nombre_inscrits_saisi)
        except (ValueError, TypeError):
            nombre_inscrits_saisi = 0

        # Calculer les suffrages exprimés pour la validation
        try:
            nombre_votants = int(request.POST.get('nombre_votants', 0) or 0)
            bulletins_nuls = int(request.POST.get('bulletins_nuls', 0) or 0)
            bulletins_blancs = int(request.POST.get('bulletins_blancs', 0) or 0)
        except (ValueError, TypeError):
            nombre_votants = bulletins_nuls = bulletins_blancs = 0

        suffrages_exprimes = nombre_votants - bulletins_nuls - bulletins_blancs

        # Créer le formset avec les données POST
        resultat_formset = ResultatFormSet(
            request.POST,
            proces_verbal=pv_existant,
            suffrages_exprimes=suffrages_exprimes
        )

        if pv_form.is_valid() and resultat_formset.is_valid():
            try:
                with transaction.atomic():
                    # Mettre à jour le nombre d'inscrits du bureau
                    bureau.nombre_inscrits = nombre_inscrits_saisi
                    bureau.save()

                    # Sauvegarder le PV
                    pv = pv_form.save(commit=False)
                    pv.bureau_vote = bureau
                    pv.representant = request.user
                    pv.save()

                    # Supprimer les anciens résultats
                    ResultatCandidat.objects.filter(proces_verbal=pv).delete()

                    # Sauvegarder les résultats des candidats
                    for i, (form, candidat) in enumerate(zip(resultat_formset.forms, candidats)):
                        if form.is_valid() and form.cleaned_data:
                            nombre_voix = form.cleaned_data.get('nombre_voix', 0)
                            if nombre_voix is None:
                                nombre_voix = 0
                            ResultatCandidat.objects.create(
                                proces_verbal=pv,
                                candidat=candidat,
                                nombre_voix=nombre_voix
                            )

                    action = "mis à jour" if pv_existant else "enregistré"
                    messages.success(
                        request,
                        f'✅ Procès-verbal {action} avec succès ! '
                        f'{nombre_inscrits_saisi} inscrits, {pv.nombre_votants} votants, {pv.suffrages_exprimes} suffrages exprimés.'
                    )
                    return redirect('saisie_resultat')

            except Exception as e:
                messages.error(request, f'Erreur lors de la sauvegarde : {str(e)}')
        else:
            # Afficher les erreurs du formulaire PV
            if not pv_form.is_valid():
                for field, errors in pv_form.errors.items():
                    for error in errors:
                        messages.error(request, f'{pv_form.fields.get(field).label if field != "__all__" else "Erreur"}: {error}')

            # Afficher les erreurs du formset
            if not resultat_formset.is_valid():
                for error in resultat_formset.non_form_errors():
                    messages.error(request, error)

                # Afficher les erreurs de chaque formulaire individuel
                for i, form in enumerate(resultat_formset.forms):
                    if form.errors:
                        candidat_name = candidats[i].get_full_name() if i < len(candidats) else f"Candidat {i+1}"
                        for field, errors in form.errors.items():
                            for error in errors:
                                messages.error(request, f'{candidat_name} - {field}: {error}')

    else:
        # GET request - Afficher le formulaire
        pv_form = ProcesVerbalForm(instance=pv_existant, bureau_vote=bureau)

        # Préparer les données initiales pour le formset
        initial_data = []
        if pv_existant:
            for candidat in candidats:
                resultat = ResultatCandidat.objects.filter(
                    proces_verbal=pv_existant,
                    candidat=candidat
                ).first()
                initial_data.append({
                    'nombre_voix': resultat.nombre_voix if resultat else 0
                })
        else:
            initial_data = [{'nombre_voix': 0} for _ in candidats]

        resultat_formset = ResultatFormSet(
            initial=initial_data,
            proces_verbal=pv_existant
        )

    # Combiner candidats avec leurs formulaires
    candidats_forms = list(zip(candidats, resultat_formset))

    context = {
        'bureau': bureau,
        'pv_form': pv_form,
        'pv_existant': pv_existant,
        'candidats_forms': candidats_forms,
        'formset': resultat_formset,
        'management_form': resultat_formset.management_form,
    }

    return render(request, 'saisie_resultat.html', context)


@login_required
def dashboard_candidat(request):
    """Tableau de bord pour un candidat - Vue de ses résultats"""
    if request.user.role != 'candidat':
        messages.error(request, 'Accès non autorisé. Cette page est réservée aux candidats.')
        return redirect('home')

    candidat = request.user

    # Récupérer tous les résultats du candidat
    resultats = ResultatCandidat.objects.filter(
        candidat=candidat
    ).select_related(
        'proces_verbal',
        'proces_verbal__bureau_vote',
        'proces_verbal__bureau_vote__centre_vote',
        'proces_verbal__bureau_vote__centre_vote__sous_prefecture',
        'proces_verbal__bureau_vote__centre_vote__sous_prefecture__departement'
    ).order_by(
        'proces_verbal__bureau_vote__centre_vote__sous_prefecture__nom',
        'proces_verbal__bureau_vote__centre_vote__nom',
        'proces_verbal__bureau_vote__numero'
    )

    # Statistiques globales
    total_voix = resultats.aggregate(Sum('nombre_voix'))['nombre_voix__sum'] or 0
    total_bureaux_avec_resultats = resultats.count()
    total_bureaux = BureauVote.objects.count()

    # Calculer le total des suffrages exprimés
    pvs = ProcesVerbal.objects.filter(
        resultats__candidat=candidat
    ).distinct()

    total_suffrages_exprimes = pvs.aggregate(Sum('suffrages_exprimes'))['suffrages_exprimes__sum'] or 0
    total_votants = pvs.aggregate(Sum('nombre_votants'))['nombre_votants__sum'] or 0
    total_inscrits = sum([pv.bureau_vote.nombre_inscrits for pv in pvs])

    # Calculs des taux
    taux_couverture = (total_bureaux_avec_resultats / total_bureaux * 100) if total_bureaux > 0 else 0
    pourcentage_voix = (total_voix / total_suffrages_exprimes * 100) if total_suffrages_exprimes > 0 else 0
    taux_participation = (total_votants / total_inscrits * 100) if total_inscrits > 0 else 0

    # Statistiques par sous-préfecture
    stats_sous_prefecture = {}
    for resultat in resultats:
        sous_pref = resultat.proces_verbal.bureau_vote.centre_vote.sous_prefecture
        if sous_pref.id not in stats_sous_prefecture:
            stats_sous_prefecture[sous_pref.id] = {
                'nom': sous_pref.nom,
                'departement': sous_pref.departement.nom,
                'total_voix': 0,
                'nombre_bureaux': 0,
                'suffrages_exprimes': 0,
            }
        stats_sous_prefecture[sous_pref.id]['total_voix'] += resultat.nombre_voix
        stats_sous_prefecture[sous_pref.id]['nombre_bureaux'] += 1
        stats_sous_prefecture[sous_pref.id]['suffrages_exprimes'] += resultat.proces_verbal.suffrages_exprimes

    # Ajouter le pourcentage par sous-préfecture
    for sp_id in stats_sous_prefecture:
        sp = stats_sous_prefecture[sp_id]
        sp['pourcentage'] = (sp['total_voix'] / sp['suffrages_exprimes'] * 100) if sp['suffrages_exprimes'] > 0 else 0

    # Statistiques par centre de vote (Top 10)
    stats_centre = {}
    for resultat in resultats:
        centre = resultat.proces_verbal.bureau_vote.centre_vote
        if centre.id not in stats_centre:
            stats_centre[centre.id] = {
                'nom': centre.nom,
                'sous_prefecture': centre.sous_prefecture.nom,
                'total_voix': 0,
                'nombre_bureaux': 0,
                'suffrages_exprimes': 0,
            }
        stats_centre[centre.id]['total_voix'] += resultat.nombre_voix
        stats_centre[centre.id]['nombre_bureaux'] += 1
        stats_centre[centre.id]['suffrages_exprimes'] += resultat.proces_verbal.suffrages_exprimes

    # Ajouter le pourcentage par centre
    for c_id in stats_centre:
        c = stats_centre[c_id]
        c['pourcentage'] = (c['total_voix'] / c['suffrages_exprimes'] * 100) if c['suffrages_exprimes'] > 0 else 0

    context = {
        'candidat': candidat,
        'resultats': resultats,
        'total_voix': total_voix,
        'total_suffrages_exprimes': total_suffrages_exprimes,
        'total_votants': total_votants,
        'total_inscrits': total_inscrits,
        'total_bureaux_avec_resultats': total_bureaux_avec_resultats,
        'total_bureaux': total_bureaux,
        'taux_couverture': round(taux_couverture, 2),
        'pourcentage_voix': round(pourcentage_voix, 2),
        'taux_participation': round(taux_participation, 2),
        'stats_sous_prefecture': sorted(stats_sous_prefecture.values(), key=lambda x: x['total_voix'], reverse=True),
        'stats_centre': sorted(stats_centre.values(), key=lambda x: x['total_voix'], reverse=True)[:10],
    }

    return render(request, 'dashboard_candidat.html', context)


@login_required
def detail_bureau(request, bureau_id):
    """Détail d'un bureau de vote"""
    if request.user.role != 'candidat':
        messages.error(request, 'Accès non autorisé.')
        return redirect('home')

    bureau = get_object_or_404(BureauVote, id=bureau_id)
    pv = ProcesVerbal.objects.filter(bureau_vote=bureau).first()

    resultats_bureau = []
    if pv:
        resultats_bureau = ResultatCandidat.objects.filter(
            proces_verbal=pv
        ).select_related('candidat').order_by('-nombre_voix')

    context = {
        'bureau': bureau,
        'pv': pv,
        'resultats_bureau': resultats_bureau,
        'candidat': request.user,
    }

    return render(request, 'detail_bureau.html', context)


# ========================================
# NOUVELLES VUES - DASHBOARD GÉNÉRAL
# ========================================

@login_required
def dashboard_general(request):
    """Dashboard général avec tous les résultats du département de Danané"""

    # Filtrer sur le département de Danané
    try:
        danane = Departement.objects.get(nom__iexact='Danané')
    except Departement.DoesNotExist:
        # Si le département n'existe pas, prendre le premier ou tous
        danane = Departement.objects.first()
        if not danane:
            messages.error(request, "Aucun département trouvé dans le système.")
            return redirect('home')

    # Statistiques globales
    total_bureaux = BureauVote.objects.filter(
        centre_vote__sous_prefecture__departement=danane
    ).count()

    bureaux_saisis = ProcesVerbal.objects.filter(
        bureau_vote__centre_vote__sous_prefecture__departement=danane
    ).count()

    bureaux_restants = total_bureaux - bureaux_saisis
    taux_saisie = (bureaux_saisis / total_bureaux * 100) if total_bureaux > 0 else 0

    # Classement des candidats
    candidats = User.objects.filter(role='candidat').annotate(
        total_voix=Sum('resultats_obtenus__nombre_voix', filter=Q(
            resultats_obtenus__proces_verbal__bureau_vote__centre_vote__sous_prefecture__departement=danane
        )),
        nombre_bureaux=Count('resultats_obtenus', filter=Q(
            resultats_obtenus__proces_verbal__bureau_vote__centre_vote__sous_prefecture__departement=danane
        ))
    ).order_by('-total_voix')

    # Calculer le total des suffrages exprimés
    total_suffrages_exprimes = ProcesVerbal.objects.filter(
        bureau_vote__centre_vote__sous_prefecture__departement=danane
    ).aggregate(Sum('suffrages_exprimes'))['suffrages_exprimes__sum'] or 0

    # Ajouter le pourcentage à chaque candidat
    classement = []
    for candidat in candidats:
        classement.append({
            'numero_candidat': candidat.numero_candidat,
            'get_full_name': candidat.get_full_name(),
            'parti_politique': candidat.parti_politique,
            'total_voix': candidat.total_voix or 0,
            'nombre_bureaux': candidat.nombre_bureaux or 0,
            'pourcentage': (candidat.total_voix / total_suffrages_exprimes * 100) if total_suffrages_exprimes > 0 and candidat.total_voix else 0
        })

    # Participation par sous-préfecture
    participation_sp = []
    for sp in SousPrefecture.objects.filter(departement=danane):
        bureaux_sp = BureauVote.objects.filter(centre_vote__sous_prefecture=sp)
        pvs_sp = ProcesVerbal.objects.filter(bureau_vote__in=bureaux_sp)

        total_inscrits = bureaux_sp.aggregate(Sum('nombre_inscrits'))['nombre_inscrits__sum'] or 0
        total_votants = pvs_sp.aggregate(Sum('nombre_votants'))['nombre_votants__sum'] or 0
        total_nuls = pvs_sp.aggregate(Sum('bulletins_nuls'))['bulletins_nuls__sum'] or 0
        total_blancs = pvs_sp.aggregate(Sum('bulletins_blancs'))['bulletins_blancs__sum'] or 0
        total_exprimes = pvs_sp.aggregate(Sum('suffrages_exprimes'))['suffrages_exprimes__sum'] or 0

        taux_participation = (total_votants / total_inscrits * 100) if total_inscrits > 0 else 0

        participation_sp.append({
            'id': sp.id,
            'nom': sp.nom,
            'total_inscrits': total_inscrits,
            'total_votants': total_votants,
            'total_nuls': total_nuls,
            'total_blancs': total_blancs,
            'total_exprimes': total_exprimes,
            'taux_participation': round(taux_participation, 2)
        })

    # Totaux départementaux
    total_inscrits = BureauVote.objects.filter(
        centre_vote__sous_prefecture__departement=danane
    ).aggregate(Sum('nombre_inscrits'))['nombre_inscrits__sum'] or 0

    total_votants = ProcesVerbal.objects.filter(
        bureau_vote__centre_vote__sous_prefecture__departement=danane
    ).aggregate(Sum('nombre_votants'))['nombre_votants__sum'] or 0

    total_nuls = ProcesVerbal.objects.filter(
        bureau_vote__centre_vote__sous_prefecture__departement=danane
    ).aggregate(Sum('bulletins_nuls'))['bulletins_nuls__sum'] or 0

    total_blancs = ProcesVerbal.objects.filter(
        bureau_vote__centre_vote__sous_prefecture__departement=danane
    ).aggregate(Sum('bulletins_blancs'))['bulletins_blancs__sum'] or 0

    taux_participation_global = (total_votants / total_inscrits * 100) if total_inscrits > 0 else 0

    context = {
        'total_bureaux': total_bureaux,
        'bureaux_saisis': bureaux_saisis,
        'bureaux_restants': bureaux_restants,
        'taux_saisie': round(taux_saisie, 2),
        'classement': classement,
        'total_suffrages_exprimes': total_suffrages_exprimes,
        'participation_sp': participation_sp,
        'total_inscrits': total_inscrits,
        'total_votants': total_votants,
        'total_nuls': total_nuls,
        'total_blancs': total_blancs,
        'taux_participation_global': round(taux_participation_global, 2),
    }

    return render(request, 'dashboard_general.html', context)


# ========================================
# EXPORTS
# ========================================

@login_required
def export_resultats_excel(request):
    """Export Excel des résultats complets"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        messages.error(request, "La bibliothèque openpyxl n'est pas installée.")
        return redirect('dashboard_general')

    # Filtrer sur Danané
    try:
        danane = Departement.objects.get(nom__iexact='Danané')
    except Departement.DoesNotExist:
        danane = Departement.objects.first()

    # Créer le workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Résultats Danané"

    # Styles
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # En-têtes
    headers = [
        'Sous-préfecture', 'Centre', 'Bureau', 'Inscrits', 'Votants',
        'Nuls', 'Blancs', 'Exprimés'
    ]

    candidats = User.objects.filter(role='candidat').order_by('numero_candidat')
    for candidat in candidats:
        headers.append(f"{candidat.get_full_name()} (N°{candidat.numero_candidat})")

    # Écrire les en-têtes
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    # Données
    row_num = 2
    for pv in ProcesVerbal.objects.select_related(
            'bureau_vote__centre_vote__sous_prefecture'
    ).filter(
        bureau_vote__centre_vote__sous_prefecture__departement=danane
    ).order_by(
        'bureau_vote__centre_vote__sous_prefecture__nom',
        'bureau_vote__centre_vote__nom',
        'bureau_vote__numero'
    ):
        row_data = [
            pv.bureau_vote.centre_vote.sous_prefecture.nom,
            pv.bureau_vote.centre_vote.nom,
            pv.bureau_vote.numero,
            pv.bureau_vote.nombre_inscrits,
            pv.nombre_votants,
            pv.bulletins_nuls,
            pv.bulletins_blancs,
            pv.suffrages_exprimes,
        ]

        for candidat in candidats:
            resultat = ResultatCandidat.objects.filter(
                proces_verbal=pv,
                candidat=candidat
            ).first()
            row_data.append(resultat.nombre_voix if resultat else 0)

        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = border
            if col_num >= 4:  # Aligner les chiffres à droite
                cell.alignment = Alignment(horizontal="right")

        row_num += 1

    # Ajuster la largeur des colonnes
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Préparer la réponse
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=resultats_danane.xlsx'
    wb.save(response)

    return response


@login_required
def export_resultats_pdf(request):
    """Export PDF des résultats - Version universelle avec ReportLab"""

    # Vérifier que ReportLab est installé
    try:
        from reportlab.lib.pagesizes import A4, letter
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
    except ImportError:
        messages.error(
            request,
            "La bibliothèque ReportLab n'est pas installée. "
            "Installez-la avec : pip install reportlab"
        )
        return redirect('dashboard_general')

    from io import BytesIO
    from datetime import datetime

    # Filtrer sur Danané
    try:
        danane = Departement.objects.get(nom__iexact='Danané')
    except Departement.DoesNotExist:
        danane = Departement.objects.first()
        if not danane:
            messages.error(request, "Aucun département trouvé.")
            return redirect('dashboard_general')

    # Récupérer les données des candidats
    candidats = User.objects.filter(role='candidat').annotate(
        total_voix=Sum('resultats_obtenus__nombre_voix', filter=Q(
            resultats_obtenus__proces_verbal__bureau_vote__centre_vote__sous_prefecture__departement=danane
        )),
        nombre_bureaux=Count('resultats_obtenus', filter=Q(
            resultats_obtenus__proces_verbal__bureau_vote__centre_vote__sous_prefecture__departement=danane
        ))
    ).order_by('-total_voix')

    # Calculer le total des suffrages exprimés
    total_suffrages_exprimes = ProcesVerbal.objects.filter(
        bureau_vote__centre_vote__sous_prefecture__departement=danane
    ).aggregate(Sum('suffrages_exprimes'))['suffrages_exprimes__sum'] or 0

    # Créer le buffer pour le PDF
    buffer = BytesIO()

    # Créer le document PDF
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=2*cm,
        leftMargin=2*cm,
        topMargin=2*cm,
        bottomMargin=2*cm,
        title=f"Résultats Électoraux - {danane.nom}",
        author="Système de Gestion Électorale"
    )

    # Container pour les éléments du PDF
    elements = []

    # Styles
    styles = getSampleStyleSheet()

    # Style pour le titre principal
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#FF8C00'),
        spaceAfter=10,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )

    # Style pour le sous-titre
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Normal'],
        fontSize=14,
        textColor=colors.HexColor('#666666'),
        spaceAfter=5,
        alignment=TA_CENTER,
        fontName='Helvetica'
    )

    # Style pour les infos
    info_style = ParagraphStyle(
        'InfoStyle',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.HexColor('#999999'),
        spaceAfter=30,
        alignment=TA_CENTER,
        fontName='Helvetica'
    )

    # Style pour le footer
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.HexColor('#999999'),
        alignment=TA_CENTER,
        fontName='Helvetica'
    )

    # Style pour les sections
    section_style = ParagraphStyle(
        'SectionTitle',
        parent=styles['Heading2'],
        fontSize=16,
        textColor=colors.HexColor('#4472C4'),
        spaceAfter=15,
        spaceBefore=20,
        alignment=TA_LEFT,
        fontName='Helvetica-Bold'
    )

    # ====== EN-TÊTE ======
    elements.append(Paragraph("🗳️ RÉSULTATS ÉLECTORAUX", title_style))
    elements.append(Paragraph(f"<b>Département de {danane.nom}</b>", subtitle_style))
    elements.append(Paragraph("Élections Législatives 2025", subtitle_style))
    elements.append(Paragraph(
        f"Document généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}",
        info_style
    ))

    # ====== SECTION CLASSEMENT ======
    elements.append(Paragraph("📊 Classement des Candidats", section_style))
    elements.append(Spacer(1, 0.5*cm))

    # Préparer les données du tableau
    data = [['Rang', 'N°', 'Candidat', 'Parti Politique', 'Voix', 'Pourcentage']]

    for i, candidat in enumerate(candidats, 1):
        # Médailles pour le top 3
        if i == 1:
            rang = '🥇'
        elif i == 2:
            rang = '🥈'
        elif i == 3:
            rang = '🥉'
        else:
            rang = str(i)

        numero = str(candidat.numero_candidat) if candidat.numero_candidat else str(i)
        nom = candidat.get_full_name()
        parti = candidat.parti_politique or "Indépendant"
        voix = f"{candidat.total_voix or 0:,}".replace(',', ' ')

        # Calculer le pourcentage
        if total_suffrages_exprimes > 0 and candidat.total_voix:
            pourcentage = f"{(candidat.total_voix / total_suffrages_exprimes * 100):.2f}%"
        else:
            pourcentage = "0.00%"

        data.append([rang, numero, nom, parti, voix, pourcentage])

    # Ligne de total
    data.append([
        '',
        '',
        '',
        'TOTAL SUFFRAGES EXPRIMÉS',
        f"{total_suffrages_exprimes:,}".replace(',', ' '),
        '100.00%'
    ])

    # Définir les largeurs de colonnes
    col_widths = [2*cm, 1.5*cm, 5*cm, 4.5*cm, 2.5*cm, 2.5*cm]

    # Créer le tableau
    table = Table(data, colWidths=col_widths, repeatRows=1)

    # Style du tableau
    table_style = TableStyle([
        # En-tête
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('TOPPADDING', (0, 0), (-1, 0), 12),

        # Corps du tableau
        ('BACKGROUND', (0, 1), (-1, -2), colors.white),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('ALIGN', (0, 1), (0, -1), 'CENTER'),  # Rang centré
        ('ALIGN', (1, 1), (1, -1), 'CENTER'),  # N° centré
        ('ALIGN', (2, 1), (2, -1), 'LEFT'),    # Nom à gauche
        ('ALIGN', (3, 1), (3, -1), 'LEFT'),    # Parti à gauche
        ('ALIGN', (4, 1), (-1, -1), 'RIGHT'),  # Voix et % à droite
        ('FONTNAME', (0, 1), (-1, -2), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -2), 10),
        ('TOPPADDING', (0, 1), (-1, -2), 8),
        ('BOTTOMPADDING', (0, 1), (-1, -2), 8),

        # Bordures
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('LINEBELOW', (0, 0), (-1, 0), 2, colors.HexColor('#4472C4')),

        # Alternance de couleurs pour les lignes
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#f9f9f9')]),

        # Ligne de total
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#e6e6e6')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, -1), (-1, -1), 11),
        ('TOPPADDING', (0, -1), (-1, -1), 10),
        ('BOTTOMPADDING', (0, -1), (-1, -1), 10),
        ('LINEABOVE', (0, -1), (-1, -1), 2, colors.HexColor('#4472C4')),
    ])

    table.setStyle(table_style)
    elements.append(table)

    # Espacement
    elements.append(Spacer(1, 1*cm))

    # ====== STATISTIQUES SUPPLÉMENTAIRES ======
    # Calculer statistiques
    total_bureaux = BureauVote.objects.filter(
        centre_vote__sous_prefecture__departement=danane
    ).count()

    bureaux_saisis = ProcesVerbal.objects.filter(
        bureau_vote__centre_vote__sous_prefecture__departement=danane
    ).count()

    taux_saisie = (bureaux_saisis / total_bureaux * 100) if total_bureaux > 0 else 0

    # Boîte d'informations
    info_data = [
        ['Statistiques de Saisie', ''],
        ['Bureaux de vote (total)', str(total_bureaux)],
        ['Procès-verbaux enregistrés', str(bureaux_saisis)],
        ['Taux de saisie', f"{taux_saisie:.1f}%"],
        ['', ''],
        ['Total suffrages exprimés', f"{total_suffrages_exprimes:,}".replace(',', ' ')],
    ]

    info_table = Table(info_data, colWidths=[8*cm, 4*cm])
    info_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('SPAN', (0, 0), (-1, 0)),
        ('TOPPADDING', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),

        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f9f9f9')),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('ALIGN', (0, 1), (0, -1), 'LEFT'),
        ('ALIGN', (1, 1), (1, -1), 'RIGHT'),
        ('TOPPADDING', (0, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),

        # Ligne vide
        ('BACKGROUND', (0, 4), (-1, 4), colors.white),
        ('GRID', (0, 4), (-1, 4), 0, colors.white),

        # Ligne total
        ('BACKGROUND', (0, 5), (-1, 5), colors.HexColor('#e6e6e6')),
        ('FONTNAME', (0, 5), (-1, 5), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 5), (-1, 5), 11),
    ]))

    elements.append(info_table)

    # ====== FOOTER ======
    elements.append(Spacer(1, 2*cm))
    elements.append(Paragraph(
        "─────────────────────────────────────────────────────────",
        footer_style
    ))
    elements.append(Spacer(1, 0.3*cm))
    elements.append(Paragraph(
        f"© 2025 Gestion des Résultats Électoraux - Département de {danane.nom}",
        footer_style
    ))
    elements.append(Paragraph(
        "Document officiel généré automatiquement",
        footer_style
    ))

    # Construire le PDF
    try:
        doc.build(elements)
    except Exception as e:
        messages.error(request, f"Erreur lors de la génération du PDF : {str(e)}")
        return redirect('dashboard_general')

    # Préparer la réponse HTTP
    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="resultats_{danane.nom.lower().replace(" ", "_")}.pdf"'

    return response


# ========================================
# API POUR LE MODAL DÉTAILS
# ========================================

@login_required
def api_sous_prefecture_bureaux(request, sous_prefecture_id):
    """API pour récupérer les détails des bureaux d'une sous-préfecture"""

    try:
        # Récupérer la sous-préfecture
        sous_prefecture = get_object_or_404(SousPrefecture, id=sous_prefecture_id)

        # Récupérer tous les centres de vote
        centres = CentreVote.objects.filter(
            sous_prefecture=sous_prefecture
        ).prefetch_related(
            'bureaux',
            'bureaux__proces_verbal'
        ).order_by('nom')

        # Construire les données JSON
        data = {
            'id': sous_prefecture.id,
            'nom': sous_prefecture.nom,
            'departement': sous_prefecture.departement.nom,
            'total_centres': centres.count(),
            'total_bureaux': 0,
            'centres': []
        }

        for centre in centres:
            bureaux = centre.bureaux.all()
            data['total_bureaux'] += bureaux.count()

            centre_data = {
                'id': centre.id,
                'nom': centre.nom,
                'adresse': centre.adresse or '',
                'bureaux': []
            }

            for bureau in bureaux:
                bureau_data = {
                    'id': bureau.id,
                    'numero': bureau.numero,
                    'nombre_inscrits': bureau.nombre_inscrits,
                    'pv': None
                }

                # Vérifier si le bureau a un PV
                try:
                    pv = bureau.proces_verbal
                    bureau_data['pv'] = {
                        'id': pv.id,
                        'nombre_votants': pv.nombre_votants,
                        'bulletins_nuls': pv.bulletins_nuls,
                        'bulletins_blancs': pv.bulletins_blancs,
                        'suffrages_exprimes': pv.suffrages_exprimes,
                        'verifie': pv.verifie,
                        'photo_pv_url': pv.photo_pv.url if pv.photo_pv else None,
                        'date_saisie': pv.date_saisie.strftime('%d/%m/%Y %H:%M') if pv.date_saisie else '',
                        'representant': pv.representant.get_full_name() if pv.representant else ''
                    }
                except ProcesVerbal.DoesNotExist:
                    pass
                except AttributeError:
                    pass

                centre_data['bureaux'].append(bureau_data)

            data['centres'].append(centre_data)

        return JsonResponse(data, safe=False)

    except SousPrefecture.DoesNotExist:
        return JsonResponse(
            {'error': 'Sous-préfecture non trouvée'},
            status=404
        )
    except Exception as e:
        # Pour le débogage, retourner l'erreur détaillée
        import traceback
        return JsonResponse(
            {
                'error': str(e),
                'traceback': traceback.format_exc()
            },
            status=500
        )